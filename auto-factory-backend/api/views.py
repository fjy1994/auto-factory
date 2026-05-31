"""
REST API 视图 — 匹配设计文档（自动化工厂设计方案.md）第6章 API接口设计
"""

from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Count, Q
from django.utils import timezone
from .models import (
    Device, Branch, Version, TaskDef, Email,
    FlashingTask, FlashingProcess,
    Task, TestCase, CaseSet, TaskResult,
)
from .serializers import (
    DeviceSerializer, BranchSerializer, VersionSerializer, TaskDefSerializer,
    EmailSerializer, FlashingTaskSerializer, FlashingProcessSerializer,
    TaskSerializer, TestCaseSerializer, CaseSetSerializer, TaskResultSerializer,
)
from .flashing_dispatcher import (
    handle_step_callback,
    retry_process_step,
    retry_process_all,
)


# ───────────────────────── Utility helpers ─────────────────────────

def ok_response(data=None, code=0):
    """成功响应包装，非查询接口统一格式"""
    result = {'code': code, 'message': 'success'}
    if data is not None:
        result['data'] = data
    return Response(result, status=status.HTTP_200_OK)


def created_response(obj_id):
    """创建成功响应"""
    return Response(
        {'code': 0, 'message': 'success', 'id': obj_id},
        status=status.HTTP_201_CREATED,
    )


def deleted_response():
    """删除成功响应（204 No Content）"""
    return Response(status=status.HTTP_204_NO_CONTENT)


# ───────────────────────── ViewSets ─────────────────────────

class DeviceViewSet(viewsets.ModelViewSet):
    """6.2 设备管理"""
    queryset = Device.objects.all()
    serializer_class = DeviceSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['serial', 'model']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        status_val = params.get('status')
        executor_ip = params.get('executor_ip')
        model = params.get('model')
        if status_val:
            qs = qs.filter(status=status_val)
        if executor_ip:
            qs = qs.filter(executor_ip__icontains=executor_ip)
        if model:
            qs = qs.filter(model__icontains=model)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return created_response(obj.id)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)  # 设计文档要求 PUT 也支持部分更新
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        # PUT 返回更新后的完整对象
        return Response(DeviceSerializer(instance).data)

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return deleted_response()


class BranchViewSet(viewsets.ModelViewSet):
    """6.3 分支管理"""
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return created_response(obj.id)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return ok_response()

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # 级联删除任务定义
        TaskDef.objects.filter(branch=instance).delete()
        instance.delete()
        return deleted_response()


class TaskDefViewSet(viewsets.ModelViewSet):
    """6.3 分支管理 — 任务定义子资源"""
    queryset = TaskDef.objects.all()
    serializer_class = TaskDefSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

    def get_queryset(self):
        qs = super().get_queryset()
        branch_id = self.request.query_params.get('branch_id')
        if branch_id:
            qs = qs.filter(branch_id=branch_id)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return created_response(obj.id)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return ok_response()

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return deleted_response()


class VersionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    6.4 版本跟踪
    ReadOnly — 版本由 LLM 邮件解析 + 逻辑模块自动创建/更新
    """
    queryset = Version.objects.all().order_by('-created_at')
    serializer_class = VersionSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['version', 'branch__name']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        branch_name = params.get('branch_name')
        version_label = params.get('version_label')
        status_val = params.get('status')
        if branch_name:
            qs = qs.filter(branch__name__icontains=branch_name)
        if version_label:
            qs = qs.filter(version__icontains=version_label)
        if status_val is not None:
            qs = qs.filter(status=int(status_val))
        return qs

    def list(self, request, *args, **kwargs):
        """重写 list：在返回数据中注入 tasks 嵌套数组"""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data

        for idx, version in enumerate(queryset):
            # tasks：关联任务的嵌套信息
            tasks_qs = Task.objects.filter(version=version)
            data[idx]['tasks'] = [
                {
                    'id': t.id,
                    'name': t.name,
                    'status': t.status,
                    'progress': t.progress,
                    'pass_count': t.pass_count,
                    'fail_count': t.fail_count,
                    'total_count': t.total_count,
                    'pass_rate': t.pass_rate,
                }
                for t in tasks_qs
            ]

        return ok_response(data)


class FlashingProcessViewSet(viewsets.ModelViewSet):
    """6.5 刷机管理"""
    queryset = FlashingProcess.objects.all()
    serializer_class = FlashingProcessSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        version_label = self.request.query_params.get('version_label')
        if version_label:
            qs = qs.filter(flashing_task__version__version__icontains=version_label)
        return qs.select_related('device', 'flashing_task', 'flashing_task__version')

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return created_response(obj.id)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return ok_response()

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return deleted_response()

    @action(detail=True, methods=['post'], url_path='retry-step')
    def retry_step(self, request, pk=None):
        """重试指定步骤（重置状态 + 通知 Agent）"""
        step_name = request.data.get('step_name')
        if not step_name:
            return Response(
                {'code': 400, 'message': 'step_name 必填'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        result = retry_process_step(pk, step_name)
        if result.get('success'):
            return ok_response()
        return Response(
            {'code': 500, 'message': result.get('error', '重试失败')},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    @action(detail=True, methods=['post'], url_path='retry-all')
    def retry_all(self, request, pk=None):
        """从第一步重新开始整个刷机流程（重置 + 通知 Agent）"""
        result = retry_process_all(pk)
        if result.get('success'):
            return ok_response()
        return Response(
            {'code': 500, 'message': result.get('error', '重试失败')},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    @action(detail=True, methods=['post'], url_path='step-result')
    def step_result(self, request, pk=None):
        """Agent 回调：上报步骤执行结果"""
        step_name = request.data.get('step_name')
        step_status = request.data.get('status')
        error_message = request.data.get('error_message', '')
        if not step_name or not step_status:
            return Response(
                {'code': 400, 'message': 'step_name 和 status 必填'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        result = handle_step_callback(pk, step_name, step_status, error_message)
        if result.get('success'):
            return ok_response(result)
        return Response(
            {'code': 500, 'message': result.get('error', '处理回调失败')},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


class FlashingTaskViewSet(viewsets.ReadOnlyModelViewSet):
    """6.x 刷机任务（版本维度）"""
    queryset = FlashingTask.objects.all().order_by('-created_at')
    serializer_class = FlashingTaskSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        version_label = self.request.query_params.get('version_label')
        if version_label:
            qs = qs.filter(version__version__icontains=version_label)
        return qs.select_related('version', 'version__branch')


class TaskViewSet(viewsets.ModelViewSet):
    """6.6 任务中心"""
    queryset = Task.objects.all()
    serializer_class = TaskSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        status_val = params.get('status')
        version_label = params.get('version_label')
        if status_val:
            qs = qs.filter(status=status_val)
        if version_label:
            qs = qs.filter(version__version__icontains=version_label)
        return qs.select_related('task_def', 'version', 'version__branch')

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return created_response(obj.id)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return ok_response()

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return deleted_response()

    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """手动开始执行任务（仅 queued 状态有效）"""
        task = self.get_object()
        if task.status != 'queued':
            return Response(
                {'code': 400, 'message': '只能启动 queued 状态的任务'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        task.status = 'running'
        task.start_time = timezone.now()
        task.save(update_fields=['status', 'start_time'])
        return ok_response()

    @action(detail=True, methods=['get'])
    def results(self, request, pk=None):
        """任务下各用例的执行结果列表"""
        task = self.get_object()
        qs = TaskResult.objects.filter(task=task)
        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        serializer = TaskResultSerializer(qs, many=True)
        return Response(serializer.data)


class TestCaseViewSet(viewsets.ModelViewSet):
    """6.7 用例管理"""
    queryset = TestCase.objects.all()
    serializer_class = TestCaseSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['case_id', 'name']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        module = params.get('module')
        priority = params.get('priority')
        keyword = params.get('keyword')
        if module:
            qs = qs.filter(module=module)
        if priority:
            qs = qs.filter(priority=priority)
        if keyword:
            qs = qs.filter(Q(case_id__icontains=keyword) | Q(name__icontains=keyword))
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return created_response(obj.id)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return ok_response()

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return deleted_response()

    @action(detail=False, methods=['post'])
    def import_cases(self, request):
        """Excel 批量导入用例"""
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'code': 400, 'message': '请上传 Excel 文件'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        import openpyxl
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response(
                {'code': 400, 'message': 'Excel 文件为空'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        headers = [str(c).lower().strip() if c else '' for c in rows[0]]
        imported = 0
        skipped = 0
        errors = []

        # 列名映射
        col_map = {
            'case_id': 'case_id',
            '用例编号': 'case_id',
            'id': 'case_id',
            'name': 'name',
            '用例名称': 'name',
            'module': 'module',
            '所属模块': 'module',
            'priority': 'priority',
            '优先级': 'priority',
            'precondition': 'precondition',
            '前置条件': 'precondition',
            '前置步骤': 'precondition',
            'steps': 'steps',
            '测试步骤': 'steps',
            'expected': 'expected',
            '预期结果': 'expected',
            'creator': 'creator',
            '责任人': 'creator',
            '创建人': 'creator',
        }

        h = {col_map.get(h, h): i for i, h in enumerate(headers)}

        def _map_priority(p):
            p = str(p).strip().lower()
            if p in ('l0', '高', 'p0', 'critical'):
                return 'L0'
            if p in ('l1', '中高', 'p1'):
                return 'L1'
            if p in ('l2', '中', 'p2', 'medium'):
                return 'L2'
            if p in ('l3', '中低', 'p3'):
                return 'L3'
            if p in ('l4', '低', 'p4', 'low'):
                return 'L4'
            return 'L2'

        for i, row in enumerate(rows[1:], start=2):
            case_id = str(row[h['case_id']]).strip() if 'case_id' in h else ''
            if not case_id:
                skipped += 1
                continue
            tc_data = {
                'case_id': case_id,
                'name': str(row[h['name']]).strip() if 'name' in h else '',
                'module': str(row[h['module']]).strip() if 'module' in h else '',
                'priority': _map_priority(row[h['priority']]) if 'priority' in h else 'L2',
                'precondition': str(row[h['precondition']]).strip() if 'precondition' in h else '',
                'steps': str(row[h['steps']]).strip() if 'steps' in h else '',
                'expected': str(row[h['expected']]).strip() if 'expected' in h else '',
                'creator': str(row[h['creator']]).strip() if 'creator' in h else '',
            }
            try:
                TestCase.objects.update_or_create(
                    case_id=case_id,
                    defaults=tc_data,
                )
                imported += 1
            except Exception as e:
                skipped += 1
                errors.append(f'第{i}行: {case_id} - {str(e)}')

        return Response({
            'imported_count': imported,
            'skipped_count': skipped,
            'errors': errors,
        })

    @action(detail=False, methods=['get'])
    def export_template(self, request):
        """下载 Excel 导入模板"""
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '测试用例模板'
        ws.append(['用例编号', '用例名称', '所属模块', '优先级', '前置步骤', '测试步骤', '预期结果', '责任人'])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=test_case_template.xlsx'
        wb.save(response)
        return response


class CaseSetViewSet(viewsets.ModelViewSet):
    """6.8 用例集管理"""
    queryset = CaseSet.objects.all()
    serializer_class = CaseSetSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return created_response(obj.id)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return ok_response()

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return deleted_response()


class EmailViewSet(viewsets.ModelViewSet):
    """邮件管理"""
    queryset = Email.objects.all()
    serializer_class = EmailSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return created_response(obj.id)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return ok_response()

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return deleted_response()


# ───────────────────────── 仪表盘 ─────────────────────────

from rest_framework.decorators import api_view


@api_view(['GET'])
def dashboard_stats(request):
    """6.1 仪表盘"""
    devices = Device.objects.all()
    versions = Version.objects.all()
    today = timezone.now().date()

    return ok_response({
        'device_stats': {
            'total': devices.count(),
            'idle': devices.filter(status='idle').count(),
            'busy': devices.filter(status='busy').count(),
            'offline': devices.filter(status='offline').count(),
        },
        'version_coverage': {
            'completed': versions.filter(status=Version.Status.COMPLETED).count(),
            'testing': versions.filter(status=Version.Status.TESTING).count(),
            'queued': versions.filter(status=Version.Status.QUEUED).count(),
            'no_device': versions.filter(status=Version.Status.NO_DEVICE).count(),
            'flash_failed': versions.filter(status=Version.Status.FLASH_FAILED).count(),
            'obsoleted': versions.filter(status=Version.Status.OBSOLETED).count(),
        },
        'recent_tasks': {
            'today_completed': Task.objects.filter(
                status='completed', end_time__date=today
            ).count(),
            'today_failed': Task.objects.filter(
                status='failed', end_time__date=today
            ).count(),
            'running': Task.objects.filter(status='running').count(),
        },
    })
